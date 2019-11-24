'''
excel写入脚本 1.0版本
'''
import django
import openpyxl
import pymysql

def mysql(database,table,ex):
    db = pymysql.connect(host='localhost',
                         port = 3306,
                         user='root',
                         password = '781124',
                         database = database,
                         charset='utf8')
    cur = db.cursor()
    # 插入全部数据
    # sql = "insert into exwork values ('%s','%s','%s','%s','%s');"
    sql = "insert into %s values (" %table

    for e in ex:
        l = e.split(',')
        # 看每行的长度 , 按长度添加'%s'
        for i in range(len(l)):
            sql += "'%s',"% l[i]
        # 去掉末尾的','
        sql = sql.rstrip(',')
        sql += ');'
        print(sql)
        cur.execute(sql)
        sql = "insert into %s values (" % table
    db.commit()
    cur.close()
    db.close()

def excel(ex):
    f = openpyxl.load_workbook(ex)
    ws = f.active
    i = 1
    txt = ''
    while True:
        wc = ws.cell(row=i, column=1)
        if not wc.value:
            break
        c = 1

        while True:
            wcc = ws.cell(row=i, column=c).value
            if not wcc:
                break
            # int转换为str
            elif type(wcc) != str:
                wcc = str(wcc)
            wcc += ','
            txt += wcc
            # print(wcc)
            c += 1

        # 去掉每行末尾的','
        txt = txt.rstrip(',')
        txt += '\n'
        i += 1

    t = txt.split('\n')
    # 去掉列表末尾的空值
    t.pop()
    # print(t)

    # 列表转成元组
    return t

# print(ws.cell(row=1, column=2)) # 获取第一行第二列的单元格
# c=ws.cell(row=1, column=49).value
# print(c)
# for i in range(1,60): #  获取1,3,4,7 行第二列的值
#     print(i, ws.cell(row=i, column=1).value)

if __name__ == '__main__':
    # 输入excel文件名x
    x = 'work13.xlsx'
    # 输入库名database
    database = 'bus'
    # 输入表名table
    table = 'exwork'
    ex = excel(x)
    mysql(database,table,ex)

    # print(ex) # ex列表
    # l = ex[0].split(',')
    # sql = "insert into exwork values ("
    # for i in range(len(l)):
    #     sql+="'%s',"
    # sql = sql.rstrip(',')
    # sql += ');'
    # print(sql)